import io

import openpyxl
from django.http import HttpResponse
from rest_framework import viewsets, permissions, filters
from rest_framework.decorators import action
from rest_framework.response import Response

from apps.users.permissions import IsAdmin
from .models import (
    Manufacturer, Vendor, InternalTeamMember,
    ManufacturerRating, VendorRating, VendorProjectPayment,
)
from .serializers import (
    ManufacturerSerializer, VendorSerializer, InternalTeamMemberSerializer,
    ManufacturerRatingSerializer, VendorRatingSerializer, VendorProjectPaymentSerializer,
)


class CRMMasterDataPermission(permissions.BasePermission):
    """All authenticated users can read and create; only admin can edit/delete."""
    def has_permission(self, request, view):
        if not request.user or not request.user.is_authenticated:
            return False
        if request.method in (*permissions.SAFE_METHODS, 'POST'):
            return True
        return request.user.role == 'admin'


class ManufacturerViewSet(viewsets.ModelViewSet):
    serializer_class = ManufacturerSerializer
    permission_classes = [CRMMasterDataPermission]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['company_name', 'poc_name', 'city', 'state']
    ordering_fields = ['company_name', 'city', 'created_at']

    def get_queryset(self):
        return Manufacturer.objects.prefetch_related('ratings').all()

    @action(detail=False, methods=['get'])
    def dropdown(self, request):
        qs = Manufacturer.objects.values('id', 'vendor_id', 'company_name', 'city').order_by('company_name')
        return Response(list(qs))

    @action(detail=False, methods=['get'], url_path='all-for-payment')
    def all_for_payment(self, request):
        """Combined list of all manufacturers + all vendors for payment vendor search."""
        manufacturers = list(
            Manufacturer.objects.values('id', 'vendor_id', 'company_name', 'city')
            .order_by('company_name')
        )
        vendors = list(
            Vendor.objects.values('id', 'vendor_id', 'company_name', 'city', 'vendor_type')
            .order_by('company_name')
        )
        result = [
            {'id': str(m['id']), 'vendor_id': m['vendor_id'], 'company_name': m['company_name'],
             'city': m['city'] or '', 'kind': 'manufacturer'}
            for m in manufacturers
        ] + [
            {'id': str(v['id']), 'vendor_id': v['vendor_id'], 'company_name': v['company_name'],
             'city': v['city'] or '', 'kind': 'vendor', 'vendor_type': v['vendor_type']}
            for v in vendors
        ]
        result.sort(key=lambda x: x['company_name'].lower())
        return Response(result)

    def destroy(self, request, *args, **kwargs):
        if request.user.role != 'admin':
            return Response({'detail': 'Only admin can delete.'}, status=403)
        return super().destroy(request, *args, **kwargs)

    # ------------------------------------------------------------------
    # GET /api/crm/manufacturers/upload-template/
    # ------------------------------------------------------------------
    @action(detail=False, methods=['get'], url_path='upload-template')
    def download_template(self, request):
        columns = [
            'company_name', 'poc_name', 'phone_no', 'email',
            'city', 'state', 'address',
            'gst_no', 'pan_no', 'bank_name', 'bank_account_no', 'bank_ifsc',
            'notes',
            'us_fda', 'cosmetic_fda', 'ayush', 'iso', 'gst_certified', 'gmp', 'stability_chamber',
        ]
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Manufacturers'
        bold_fill = openpyxl.styles.PatternFill('solid', fgColor='FFF3CD')
        bold_font = openpyxl.styles.Font(bold=True)
        hint_font = openpyxl.styles.Font(italic=True, color='888888')

        for i, col in enumerate(columns, 1):
            c = ws.cell(row=1, column=i, value=col)
            c.font = bold_font
            c.fill = bold_fill

        widths = [30, 20, 15, 25, 15, 15, 30, 18, 15, 20, 20, 15, 30, 10, 14, 10, 10, 14, 10, 20]
        for i, w in enumerate(widths[:len(columns)], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        bool_hint = 'yes / no'
        for col_name in ('us_fda', 'cosmetic_fda', 'ayush', 'iso', 'gst_certified', 'gmp', 'stability_chamber'):
            ws.cell(row=2, column=columns.index(col_name) + 1, value=bool_hint).font = hint_font
        ws.cell(row=2, column=1, value='Required').font = hint_font

        # Example row
        ws.cell(row=3, column=1, value='Acme Pharma Pvt Ltd')
        ws.cell(row=3, column=2, value='Ravi Kumar')
        ws.cell(row=3, column=3, value='9876543210')
        ws.cell(row=3, column=4, value='ravi@acmepharma.com')
        ws.cell(row=3, column=5, value='Mumbai')
        ws.cell(row=3, column=6, value='Maharashtra')
        ws.cell(row=3, column=13, value='ISO and GMP certified')
        ws.cell(row=3, column=14, value='no')
        ws.cell(row=3, column=15, value='yes')
        ws.cell(row=3, column=17, value='yes')
        ws.cell(row=3, column=19, value='yes')

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        res = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        res['Content-Disposition'] = 'attachment; filename="manufacturer_upload_template.xlsx"'
        return res

    # ------------------------------------------------------------------
    # POST /api/crm/manufacturers/bulk-upload/
    # ------------------------------------------------------------------
    @action(detail=False, methods=['post'], url_path='bulk-upload')
    def bulk_upload(self, request):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'detail': 'No file provided. Send an Excel file as "file".'}, status=400)
        try:
            wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
        except Exception:
            return Response({'detail': 'Could not read file. Upload a valid .xlsx file.'}, status=400)

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return Response({'detail': 'File is empty.'}, status=400)

        header = [str(c).strip().lower() if c is not None else '' for c in rows[0]]

        def idx(name):
            try: return header.index(name)
            except ValueError: return None

        def cell(row, i):
            if i is None or i >= len(row): return None
            v = row[i]
            return str(v).strip() if v is not None else None

        def parse_bool(row, i):
            v = cell(row, i)
            return str(v).lower() in ('yes', 'y', 'true', '1') if v else False

        col = {n: idx(n) for n in [
            'company_name', 'poc_name', 'phone_no', 'email', 'city', 'state', 'address',
            'gst_no', 'pan_no', 'bank_name', 'bank_account_no', 'bank_ifsc', 'notes',
            'us_fda', 'cosmetic_fda', 'ayush', 'iso', 'gst_certified', 'gmp', 'stability_chamber',
        ]}

        if col['company_name'] is None:
            return Response({'detail': 'Column "company_name" not found in header row.'}, status=400)

        created, skipped = [], []
        pending = []
        seen_names = set()

        for row_num, row in enumerate(rows[1:], start=2):
            if all(v is None or str(v).strip() == '' for v in row):
                continue
            name = cell(row, col['company_name'])
            if not name:
                skipped.append({'row': row_num, 'company_name': '—', 'reason': 'Missing company_name'})
                continue
            name_lower = name.lower()
            if name_lower in seen_names:
                skipped.append({'row': row_num, 'company_name': name, 'reason': 'Duplicate within uploaded file'})
                continue
            seen_names.add(name_lower)
            pending.append((row_num, name, row))

        from django.db.models.functions import Lower
        existing_names_lower = set(
            Manufacturer.objects.annotate(name_lower=Lower('company_name'))
            .filter(name_lower__in=[p[1].lower() for p in pending])
            .values_list('name_lower', flat=True)
        ) if pending else set()

        for row_num, name, row in pending:
            if name.lower() in existing_names_lower:
                skipped.append({'row': row_num, 'company_name': name, 'reason': 'Company name already exists'})
                continue
            m = Manufacturer(
                company_name=name,
                poc_name=cell(row, col['poc_name']) or '',
                phone_no=cell(row, col['phone_no']) or '',
                email=cell(row, col['email']) or '',
                city=cell(row, col['city']) or '',
                state=cell(row, col['state']) or '',
                address=cell(row, col['address']) or '',
                gst_no=cell(row, col['gst_no']) or '',
                pan_no=cell(row, col['pan_no']) or '',
                bank_name=cell(row, col['bank_name']) or '',
                bank_account_no=cell(row, col['bank_account_no']) or '',
                bank_ifsc=cell(row, col['bank_ifsc']) or '',
                notes=cell(row, col['notes']) or '',
                us_fda=parse_bool(row, col['us_fda']),
                cosmetic_fda=parse_bool(row, col['cosmetic_fda']),
                ayush=parse_bool(row, col['ayush']),
                iso=parse_bool(row, col['iso']),
                gst_certified=parse_bool(row, col['gst_certified']),
                gmp=parse_bool(row, col['gmp']),
                stability_chamber=parse_bool(row, col['stability_chamber']),
            )
            m.save()  # generates vendor_id via model.save()
            created.append({'row': row_num, 'company_name': name, 'vendor_id': m.vendor_id})

        return Response({'created': created, 'skipped': skipped}, status=200)


class VendorViewSet(viewsets.ModelViewSet):
    serializer_class = VendorSerializer
    permission_classes = [CRMMasterDataPermission]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['company_name', 'poc_name', 'city']
    ordering_fields = ['company_name', 'city', 'created_at']

    def get_queryset(self):
        qs = Vendor.objects.prefetch_related('ratings').all()
        vendor_type = self.request.query_params.get('vendor_type')
        if vendor_type:
            qs = qs.filter(vendor_type=vendor_type)
        return qs

    @action(detail=False, methods=['get'])
    def dropdown(self, request):
        vendor_type = request.query_params.get('vendor_type', '')
        qs = Vendor.objects.filter(vendor_type=vendor_type).values(
            'id', 'vendor_id', 'company_name', 'city'
        ).order_by('company_name')
        return Response(list(qs))

    def destroy(self, request, *args, **kwargs):
        if request.user.role != 'admin':
            return Response({'detail': 'Only admin can delete.'}, status=403)
        return super().destroy(request, *args, **kwargs)

    # ------------------------------------------------------------------
    # GET /api/crm/vendors/upload-template/?vendor_type=packaging
    # ------------------------------------------------------------------
    @action(detail=False, methods=['get'], url_path='upload-template')
    def download_template(self, request):
        vendor_type = request.query_params.get('vendor_type', 'vendor')
        columns = [
            'company_name', 'poc_name', 'phone_no', 'email', 'city',
            'gst_no', 'pan_no', 'bank_name', 'bank_account_no', 'bank_ifsc', 'notes',
        ]
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f'{vendor_type.title()} Vendors'
        bold_fill = openpyxl.styles.PatternFill('solid', fgColor='FFF3CD')
        bold_font = openpyxl.styles.Font(bold=True)
        hint_font = openpyxl.styles.Font(italic=True, color='888888')

        for i, col in enumerate(columns, 1):
            c = ws.cell(row=1, column=i, value=col)
            c.font = bold_font
            c.fill = bold_fill

        widths = [30, 20, 15, 25, 15, 18, 15, 20, 20, 15, 30]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        ws.cell(row=2, column=1, value='Required').font = hint_font
        ws.cell(row=2, column=2, value=f'vendor_type will be set to "{vendor_type}" automatically').font = hint_font

        ws.cell(row=3, column=1, value='Sunrise Packaging Pvt Ltd')
        ws.cell(row=3, column=2, value='Priya Sharma')
        ws.cell(row=3, column=3, value='9988776655')
        ws.cell(row=3, column=4, value='priya@sunrisepkg.com')
        ws.cell(row=3, column=5, value='Delhi')

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        res = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        res['Content-Disposition'] = f'attachment; filename="{vendor_type}_vendor_upload_template.xlsx"'
        return res

    # ------------------------------------------------------------------
    # POST /api/crm/vendors/bulk-upload/?vendor_type=packaging
    # ------------------------------------------------------------------
    @action(detail=False, methods=['post'], url_path='bulk-upload')
    def bulk_upload(self, request):
        vendor_type = request.query_params.get('vendor_type', '')
        valid_types = [t[0] for t in Vendor.VENDOR_TYPES]
        if vendor_type not in valid_types:
            return Response({'detail': f'Invalid vendor_type. Must be one of: {valid_types}'}, status=400)

        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'detail': 'No file provided. Send an Excel file as "file".'}, status=400)
        try:
            wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
        except Exception:
            return Response({'detail': 'Could not read file. Upload a valid .xlsx file.'}, status=400)

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return Response({'detail': 'File is empty.'}, status=400)

        header = [str(c).strip().lower() if c is not None else '' for c in rows[0]]

        def idx(name):
            try: return header.index(name)
            except ValueError: return None

        def cell(row, i):
            if i is None or i >= len(row): return None
            v = row[i]
            return str(v).strip() if v is not None else None

        col = {n: idx(n) for n in [
            'company_name', 'poc_name', 'phone_no', 'email', 'city',
            'gst_no', 'pan_no', 'bank_name', 'bank_account_no', 'bank_ifsc', 'notes',
        ]}

        if col['company_name'] is None:
            return Response({'detail': 'Column "company_name" not found in header row.'}, status=400)

        created, skipped = [], []
        pending = []
        seen_names = set()

        for row_num, row in enumerate(rows[1:], start=2):
            if all(v is None or str(v).strip() == '' for v in row):
                continue
            name = cell(row, col['company_name'])
            if not name:
                skipped.append({'row': row_num, 'company_name': '—', 'reason': 'Missing company_name'})
                continue
            name_lower = name.lower()
            if name_lower in seen_names:
                skipped.append({'row': row_num, 'company_name': name, 'reason': 'Duplicate within uploaded file'})
                continue
            seen_names.add(name_lower)
            pending.append((row_num, name, row))

        from django.db.models.functions import Lower
        existing_names_lower = set(
            Vendor.objects.filter(vendor_type=vendor_type)
            .annotate(name_lower=Lower('company_name'))
            .filter(name_lower__in=[p[1].lower() for p in pending])
            .values_list('name_lower', flat=True)
        ) if pending else set()

        for row_num, name, row in pending:
            if name.lower() in existing_names_lower:
                skipped.append({'row': row_num, 'company_name': name, 'reason': 'Company name already exists'})
                continue
            v = Vendor(
                vendor_type=vendor_type,
                company_name=name,
                poc_name=cell(row, col['poc_name']) or '',
                phone_no=cell(row, col['phone_no']) or '',
                email=cell(row, col['email']) or '',
                city=cell(row, col['city']) or '',
                gst_no=cell(row, col['gst_no']) or '',
                pan_no=cell(row, col['pan_no']) or '',
                bank_name=cell(row, col['bank_name']) or '',
                bank_account_no=cell(row, col['bank_account_no']) or '',
                bank_ifsc=cell(row, col['bank_ifsc']) or '',
                notes=cell(row, col['notes']) or '',
            )
            v.save()  # generates vendor_id via model.save()
            created.append({'row': row_num, 'company_name': name, 'vendor_id': v.vendor_id})

        return Response({'created': created, 'skipped': skipped}, status=200)


class InternalTeamMemberViewSet(viewsets.ModelViewSet):
    serializer_class = InternalTeamMemberSerializer
    permission_classes = [CRMMasterDataPermission]
    filter_backends = [filters.SearchFilter]
    search_fields = ['name', 'email']

    def get_queryset(self):
        qs = InternalTeamMember.objects.select_related('user').all()
        team = self.request.query_params.get('team')
        if team:
            qs = qs.filter(team=team)
        return qs

    def destroy(self, request, *args, **kwargs):
        if request.user.role != 'admin':
            return Response({'detail': 'Only admin can delete.'}, status=403)
        return super().destroy(request, *args, **kwargs)


class ManufacturerRatingViewSet(viewsets.ModelViewSet):
    serializer_class = ManufacturerRatingSerializer
    permission_classes = [permissions.IsAuthenticated]
    http_method_names = ['get', 'post', 'head', 'options']

    def get_queryset(self):
        return ManufacturerRating.objects.select_related('rated_by').all()

    def perform_create(self, serializer):
        serializer.save(rated_by=self.request.user)


class VendorRatingViewSet(viewsets.ModelViewSet):
    serializer_class = VendorRatingSerializer
    permission_classes = [permissions.IsAuthenticated]
    http_method_names = ['get', 'post', 'head', 'options']

    def get_queryset(self):
        return VendorRating.objects.select_related('rated_by').all()

    def perform_create(self, serializer):
        serializer.save(rated_by=self.request.user)


class VendorProjectPaymentViewSet(viewsets.ModelViewSet):
    serializer_class = VendorProjectPaymentSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        qs = VendorProjectPayment.objects.select_related('manufacturer', 'vendor', 'project')
        project_id = self.request.query_params.get('project')
        if project_id:
            qs = qs.filter(project_id=project_id)
        return qs

    def destroy(self, request, *args, **kwargs):
        if request.user.role != 'admin':
            return Response({'detail': 'Only admin can delete.'}, status=403)
        return super().destroy(request, *args, **kwargs)
