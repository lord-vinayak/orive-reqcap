import io
from datetime import date, datetime

import openpyxl
from django.db.models import Q
from django.http import HttpResponse
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from .models import BatchRecord
from .serializers import BatchRecordSerializer

TEMPLATE_COLUMNS = [
    'client_name', 'brand_name', 'product_type', 'product_name',
    'packaging_type', 'pack_size', 'moq', 'batch_number',
    'manufacturing_date', 'expiry_date',
]

_DATE_FORMATS = [
    '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y',
    '%b %Y', '%B %Y', '%b-%Y', '%B-%Y',  # month + year only, e.g. "Feb 2026" — defaults to the 1st
]


def _parse_date(raw):
    if raw is None or raw == '':
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    raw = str(raw).strip()
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


class BatchRecordViewSet(viewsets.ModelViewSet):
    serializer_class = BatchRecordSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = BatchRecord.objects.all()
        q = self.request.query_params.get('q')
        if q:
            qs = qs.filter(
                Q(client_name__icontains=q) | Q(brand_name__icontains=q) |
                Q(product_name__icontains=q) | Q(batch_number__icontains=q)
            )
        return qs

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    def destroy(self, request, *args, **kwargs):
        if request.user.role != 'admin':
            return Response({'detail': 'Only admin can delete.'}, status=403)
        return super().destroy(request, *args, **kwargs)

    # ------------------------------------------------------------------
    # GET /api/batch-records/upload-template/
    # ------------------------------------------------------------------
    @action(detail=False, methods=['get'], url_path='upload-template')
    def download_template(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Batch Register'

        header_font = openpyxl.styles.Font(bold=True)
        header_fill = openpyxl.styles.PatternFill('solid', fgColor='FFF3CD')

        for col_idx, col_name in enumerate(TEMPLATE_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill

        widths = [22, 22, 18, 26, 18, 12, 14, 16, 18, 14]
        for col_idx, width in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        hint_font = openpyxl.styles.Font(italic=True, color='888888')
        date_col = TEMPLATE_COLUMNS.index('manufacturing_date') + 1
        ws.cell(row=2, column=date_col, value='YYYY-MM-DD, DD-MM-YYYY, or month + year e.g. Feb 2026').font = hint_font
        exp_col = TEMPLATE_COLUMNS.index('expiry_date') + 1
        ws.cell(row=2, column=exp_col, value='YYYY-MM-DD, DD-MM-YYYY, or month + year e.g. Jan 2028').font = hint_font

        ws.cell(row=3, column=1, value='Acme Corp')
        ws.cell(row=3, column=2, value='Acme Glow')
        ws.cell(row=3, column=3, value='Face Serum')
        ws.cell(row=3, column=4, value='Vitamin C Serum')
        ws.cell(row=3, column=5, value='Glass Bottle')
        ws.cell(row=3, column=6, value='30ml')
        ws.cell(row=3, column=7, value=1000)
        ws.cell(row=3, column=8, value='B2026001')
        ws.cell(row=3, column=9, value='2026-06-01')
        ws.cell(row=3, column=10, value='2028-06-01')

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="batch_register_template.xlsx"'
        return response

    # ------------------------------------------------------------------
    # POST /api/batch-records/bulk-upload/
    # ------------------------------------------------------------------
    @action(detail=False, methods=['post'], url_path='bulk-upload')
    def bulk_upload(self, request):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'detail': 'No file provided. Send an Excel file as "file".'}, status=400)

        try:
            wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
        except Exception:
            return Response({'detail': 'Could not read the file. Please upload a valid .xlsx file.'}, status=400)

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return Response({'detail': 'The file is empty.'}, status=400)

        header_row = [str(c).strip().lower() if c is not None else '' for c in rows[0]]

        def col(name):
            try:
                return header_row.index(name)
            except ValueError:
                return None

        idx = {name: col(name) for name in TEMPLATE_COLUMNS}

        if idx['client_name'] is None or idx['product_name'] is None or idx['batch_number'] is None:
            return Response(
                {'detail': 'Required columns "client_name", "product_name" and "batch_number" not found in the header row.'},
                status=400,
            )

        def cell(row_data, name):
            i = idx[name]
            if i is None or i >= len(row_data):
                return None
            v = row_data[i]
            return str(v).strip() if v is not None else None

        def int_or_none(row_data, name):
            v = cell(row_data, name)
            if v is None:
                return None
            try:
                return int(float(v))
            except (ValueError, TypeError):
                return None

        created = []
        skipped = []
        to_create = []

        for row_num, row_data in enumerate(rows[1:], start=2):
            if all(v is None or str(v).strip() == '' for v in row_data):
                continue

            client_name = cell(row_data, 'client_name')
            product_name = cell(row_data, 'product_name')
            batch_number = cell(row_data, 'batch_number')

            if not client_name or not product_name or not batch_number:
                skipped.append({
                    'row': row_num,
                    'client_name': client_name or '—',
                    'batch_number': batch_number or '—',
                    'reason': 'Missing client_name, product_name, or batch_number',
                })
                continue

            mfg_raw = row_data[idx['manufacturing_date']] if idx['manufacturing_date'] is not None and idx['manufacturing_date'] < len(row_data) else None
            exp_raw = row_data[idx['expiry_date']] if idx['expiry_date'] is not None and idx['expiry_date'] < len(row_data) else None
            mfg_date = _parse_date(mfg_raw)
            exp_date = _parse_date(exp_raw)

            to_create.append(BatchRecord(
                client_name=client_name,
                brand_name=cell(row_data, 'brand_name') or '',
                product_type=cell(row_data, 'product_type') or '',
                product_name=product_name,
                packaging_type=cell(row_data, 'packaging_type') or '',
                pack_size=cell(row_data, 'pack_size') or '',
                moq=int_or_none(row_data, 'moq'),
                batch_number=batch_number,
                manufacturing_date=mfg_date,
                expiry_date=exp_date,
                created_by=request.user,
            ))
            entry = {'row': row_num, 'client_name': client_name, 'batch_number': batch_number}
            unparsed = []
            if mfg_raw and not mfg_date:
                unparsed.append(f'manufacturing_date "{mfg_raw}"')
            if exp_raw and not exp_date:
                unparsed.append(f'expiry_date "{exp_raw}"')
            if unparsed:
                entry['warning'] = f'Could not parse {", ".join(unparsed)} — left blank'
            created.append(entry)

        BatchRecord.objects.bulk_create(to_create)
        wb.close()
        return Response({'created': created, 'skipped': skipped}, status=200)
