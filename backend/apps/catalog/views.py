import io
from datetime import date

from django.db.models import Q
from django.http import HttpResponse
from rest_framework import viewsets, status
from rest_framework.decorators import action, api_view, parser_classes, permission_classes
from rest_framework.parsers import MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from .models import CatalogItem
from .serializers import CatalogItemSerializer
from apps.users.permissions import IsAdminOrReadOnly

# ---------------------------------------------------------------------------
# Dropdown option lists (kept in sync with frontend dropdownOptions.ts)
# ---------------------------------------------------------------------------
_BODY_PARTS = ['Face', 'Body', 'Hair', 'Lip', 'Eye']
_PRODUCT_TYPES = [
    'Wash', 'Moisturizer', 'Serum', 'Toner', 'Mask',
    'Sunscreen', 'Scrub', 'Oil', 'Shampoo', 'Conditioner', 'Spray', 'Balm',
]
_SUB_PRODUCT_TYPES = sorted({
    'Foaming', 'Gel', 'Creamy', 'Gel with beads', 'Pearly', 'Transparent', 'Gloss',
    'Gel Cream', 'Cream - light', 'Cream - thick', 'Lotion', 'Balm',
    'Water Based', 'Oil Based', 'Water', 'Milky', 'Clay', 'Cream',
    'Spray', 'Physical', 'Chemical', 'Light Weight', 'Normal', 'Mist', 'Deodorant',
})
_KEY_BENEFITS = sorted({
    'Acne', 'Pigmentation', 'Glow', 'Skin Lightening', 'Aging', 'Wrinkle',
    'Hydration', 'Barrier Repair', 'Dark Circle', 'Sun Protection',
    'Dandruff', 'Thinning', 'Greying', 'Growth', 'Frizz', 'Bond Repair',
})
_SIZES = ['8', '15', '30', '50', '100', '150', '200']
_PACKAGING_TYPES = ['Jar', 'Bottle', 'Tube', 'Stick']
_RATE_CATEGORIES = ['Basic', 'Premium', 'Luxury']
_YES_NO = ['Yes', 'No']

# Column map: Excel header (lowercased) → model field name
_COLUMN_MAP = {
    'date': 'date', 'poc': 'poc', 'client name': 'client_name',
    'body part': 'body_part', 'product type': 'product_type',
    'sub product type': 'sub_product_type',
    'kb tag1': 'kb_tag1', 'kb tag2': 'kb_tag2', 'kb tag3': 'kb_tag3',
    'specific ingredients': 'specific_ingredients',
    'color': 'color', 'fragrance': 'fragrance',
    'size': 'size', 'packaging type': 'packaging_type',
    'rate category': 'rate_category',
    'per kg rate': 'per_kg_rate', 'manufacturing cost': 'manufacturing_cost',
    'rate per unit': 'rate_per_unit',
    'tentative packaging cost': 'tentative_packaging_cost',
    'label cost': 'label_cost',
    'tentative monocarton cost': 'tentative_monocarton_cost',
    'total cost': 'total_cost', 'potential mrp': 'potential_mrp',
}
_DECIMAL_FIELDS = {
    'per_kg_rate', 'manufacturing_cost', 'rate_per_unit',
    'tentative_packaging_cost', 'label_cost',
    'tentative_monocarton_cost', 'total_cost', 'potential_mrp',
}


class CatalogViewSet(viewsets.ModelViewSet):
    serializer_class = CatalogItemSerializer
    permission_classes = [IsAuthenticated, IsAdminOrReadOnly]

    def get_queryset(self):
        qs = CatalogItem.objects.filter(is_active=True)
        params = self.request.query_params

        body_part = params.get('body_part')
        product_type = params.get('product_type')
        sub_product_type = params.get('sub_product_type')
        key_benefits = params.getlist('key_benefit')
        rate_category = params.get('rate_category')
        q = params.get('q')

        if body_part:
            qs = qs.filter(body_part__iexact=body_part)
        if product_type:
            qs = qs.filter(product_type__iexact=product_type)
        if sub_product_type:
            qs = qs.filter(sub_product_type__iexact=sub_product_type)
        if key_benefits:
            kb_filter = Q()
            for kb in key_benefits:
                kb_filter |= (
                    Q(kb_tag1__iexact=kb) |
                    Q(kb_tag2__iexact=kb) |
                    Q(kb_tag3__iexact=kb)
                )
            qs = qs.filter(kb_filter)
        if rate_category:
            qs = qs.filter(rate_category__iexact=rate_category)
        if q:
            qs = qs.filter(
                Q(body_part__icontains=q) |
                Q(product_type__icontains=q) |
                Q(sub_product_type__icontains=q) |
                Q(kb_tag1__icontains=q) |
                Q(kb_tag2__icontains=q) |
                Q(kb_tag3__icontains=q) |
                Q(specific_ingredients__icontains=q) |
                Q(client_name__icontains=q)
            )
        return qs

    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser])
    def import_xlsx(self, request):
        """Admin only: upload .xlsx file to append catalog rows (never deletes existing)."""
        if request.user.role != 'admin':
            return Response({'detail': 'Admin only'}, status=status.HTTP_403_FORBIDDEN)
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'detail': 'No file uploaded'}, status=status.HTTP_400_BAD_REQUEST)

        from openpyxl import load_workbook
        wb = load_workbook(file_obj, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return Response({'detail': 'Sheet is empty'}, status=status.HTTP_400_BAD_REQUEST)

        # Detect header row (row 1 or 2 depending on template style)
        header_row_idx = 1 if rows[0][0] is None else 0
        header = [str(h).strip() if h is not None else '' for h in rows[header_row_idx]]
        data_rows = rows[header_row_idx + 1:]

        # Stamp all new rows with today's upload date
        today = date.today()

        created = 0
        for raw in data_rows:
            if not raw or all(v is None for v in raw):
                continue
            obj = {'uploaded_at': today}
            for i, value in enumerate(raw):
                if i >= len(header):
                    break
                col = header[i].lower().strip()
                field = _COLUMN_MAP.get(col)
                if not field:
                    continue
                if value is None:
                    continue
                if field in _DECIMAL_FIELDS:
                    try:
                        obj[field] = float(value)
                    except (ValueError, TypeError):
                        continue
                else:
                    obj[field] = str(value).strip()
            if len(obj) > 1:  # more than just uploaded_at
                CatalogItem.objects.create(**obj)
                created += 1

        return Response({'created': created, 'appended': True})

    @action(detail=False, methods=['get'], url_path='download_template')
    def download_template(self, request):
        """Admin only: download a pre-formatted .xlsx template with dropdown validation."""
        if request.user.role != 'admin':
            return Response({'detail': 'Admin only'}, status=status.HTTP_403_FORBIDDEN)

        xlsx_data = _build_catalog_template()
        response = HttpResponse(
            xlsx_data,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="catalog_template.xlsx"'
        return response


def _build_catalog_template() -> bytes:
    """Build an .xlsx catalog template with native Excel dropdown validation."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = 'Catalog'

    # ── Hidden "Lists" sheet holds long dropdown values ────────────────────
    ws_lists = wb.create_sheet('Lists')

    def _write_list(col, values):
        for row_idx, val in enumerate(values, 1):
            ws_lists.cell(row=row_idx, column=col, value=val)
        col_letter = ws_lists.cell(row=1, column=col).column_letter
        return f'Lists!${col_letter}$1:${col_letter}${len(values)}'

    sub_range = _write_list(1, _SUB_PRODUCT_TYPES)
    kb_range  = _write_list(2, _KEY_BENEFITS)

    ws_lists.sheet_state = 'hidden'

    # ── Column headers ────────────────────────────────────────────────────
    headers = [
        'Date', 'POC', 'Client Name',
        'Body Part', 'Product Type', 'Sub Product Type',
        'KB Tag1', 'KB Tag2', 'KB Tag3',
        'Specific Ingredients',
        'Color', 'Fragrance',
        'Size', 'Packaging Type', 'Rate Category',
        'Per KG Rate', 'Manufacturing Cost', 'Rate Per Unit',
        'Tentative Packaging Cost', 'Label Cost',
        'Tentative Monocarton Cost', 'Total Cost', 'Potential MRP',
    ]

    header_fill = PatternFill('solid', fgColor='F5C842')
    header_font = Font(bold=True)

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Freeze the header row
    ws.freeze_panes = 'A2'

    # ── Helper to add a data validation ──────────────────────────────────
    DATA_ROWS = 1000

    def _add_inline_dv(col_letter, formula1):
        dv = DataValidation(
            type='list',
            formula1=formula1,
            allow_blank=True,
            showDropDown=False,
            showErrorMessage=True,
            error='Value not in the allowed list.',
            errorTitle='Invalid value',
        )
        ws.add_data_validation(dv)
        dv.sqref = f'{col_letter}2:{col_letter}{DATA_ROWS}'

    def _add_range_dv(col_letter, sheet_range):
        dv = DataValidation(
            type='list',
            formula1=sheet_range,
            allow_blank=True,
            showDropDown=False,
            showErrorMessage=True,
            error='Value not in the allowed list.',
            errorTitle='Invalid value',
        )
        ws.add_data_validation(dv)
        dv.sqref = f'{col_letter}2:{col_letter}{DATA_ROWS}'

    # ── Apply validations (column letters match header order above) ───────
    # D = Body Part
    _add_inline_dv('D', '"' + ','.join(_BODY_PARTS) + '"')
    # E = Product Type
    _add_inline_dv('E', '"' + ','.join(_PRODUCT_TYPES) + '"')
    # F = Sub Product Type  (too long for inline — use Lists sheet)
    _add_range_dv('F', sub_range)
    # G/H/I = KB Tag1/2/3
    _add_range_dv('G', kb_range)
    _add_range_dv('H', kb_range)
    _add_range_dv('I', kb_range)
    # K = Color
    _add_inline_dv('K', '"Yes,No"')
    # L = Fragrance
    _add_inline_dv('L', '"Yes,No"')
    # M = Size
    _add_inline_dv('M', '"' + ','.join(_SIZES) + '"')
    # N = Packaging Type
    _add_inline_dv('N', '"' + ','.join(_PACKAGING_TYPES) + '"')
    # O = Rate Category
    _add_inline_dv('O', '"' + ','.join(_RATE_CATEGORIES) + '"')

    # ── Set column widths ─────────────────────────────────────────────────
    widths = [12, 15, 20, 12, 15, 18, 16, 16, 16, 25, 8, 10, 8, 15, 14,
              14, 18, 14, 24, 12, 24, 12, 14]
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Facets endpoint
# ---------------------------------------------------------------------------
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def catalog_facets(request):
    """Return distinct values per dimension for client-side filtering UI."""
    qs = CatalogItem.objects.filter(is_active=True)
    body_part = request.query_params.get('body_part')
    product_type = request.query_params.get('product_type')

    bp_q = qs
    pt_q = qs.filter(body_part__iexact=body_part) if body_part else qs
    sp_q = pt_q.filter(product_type__iexact=product_type) if product_type else pt_q

    def _distinct(qs, field):
        return sorted({v for v in qs.values_list(field, flat=True) if v})

    return Response({
        'body_parts': _distinct(bp_q, 'body_part'),
        'product_types': _distinct(pt_q, 'product_type'),
        'sub_product_types': _distinct(sp_q, 'sub_product_type'),
        'key_benefits': sorted({
            v for v in (
                list(qs.values_list('kb_tag1', flat=True)) +
                list(qs.values_list('kb_tag2', flat=True)) +
                list(qs.values_list('kb_tag3', flat=True))
            ) if v
        }),
        'rate_categories': _distinct(qs, 'rate_category'),
    })
