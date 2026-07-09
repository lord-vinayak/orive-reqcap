import io
import json
import re

from django.utils import timezone

import openpyxl
from django.core.mail import EmailMultiAlternatives
from django.db.models import Q
from django.http import HttpResponse
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from .models import Client, EmailLog
from .serializers import ClientSerializer, EmailLogSerializer
from . import welcome_email_template as welcome_tpl
from . import reminder_email_template_1 as reminder1_tpl
from . import reminder_email_template_2 as reminder2_tpl
from . import closure_email_template as closure_tpl
from . import sample_initiation_email_template as sample_initiation_tpl
from . import sample_payment_confirmation_email_template as sample_payment_tpl
from . import sample_approval_email_template as sample_approval_tpl
from . import order_initiation_email_template as order_initiation_tpl
from . import packaging_confirmation_email_template as packaging_confirmation_tpl
from . import packaging_payment_confirmation_email_template as packaging_payment_tpl
from . import printing_confirmation_email_template as printing_confirmation_tpl
from . import printing_payment_confirmation_email_template as printing_payment_tpl
from . import final_order_shipment_email_template as final_order_shipment_tpl
from . import invoice_email_template as invoice_tpl
from . import payment_confirmation_email_template as payment_confirmation_tpl
from . import order_confirmation_email_template as order_confirmation_tpl

LEAD_BUCKETS = {
    'prospective':      ['initial_conversation', 'proposal', 'costing'],
    'sample':           ['sample'],
    'converted':        ['order'],
    'business_earned':  ['production', 'testing', 'filling', 'order_dispatch', 'order_closed'],
    'lost':             ['lead_closed'],
}

_TEMPLATE_MAP = {
    'welcome': welcome_tpl,
    'reminder_1': reminder1_tpl,
    'reminder_2': reminder2_tpl,
    'closure': closure_tpl,
    'sample_initiation': sample_initiation_tpl,
    'sample_payment_confirmation': sample_payment_tpl,
    'sample_approval': sample_approval_tpl,
    'order_initiation': order_initiation_tpl,
    'packaging_confirmation': packaging_confirmation_tpl,
    'packaging_payment_confirmation': packaging_payment_tpl,
    'printing_confirmation': printing_confirmation_tpl,
    'printing_payment_confirmation': printing_payment_tpl,
    'final_order_shipment': final_order_shipment_tpl,
    'invoice': invoice_tpl,
    'payment_confirmation': payment_confirmation_tpl,
    'order_confirmation': order_confirmation_tpl,
}

_TEMPLATE_LABELS = {
    'welcome': 'Welcome Email',
    'reminder_1': 'Reminder Email 1',
    'reminder_2': 'Reminder Email 2',
    'closure': 'Closure Email',
    'sample_initiation': 'Sample Initiation Email',
    'sample_payment_confirmation': 'Sample Payment Confirmation Email',
    'sample_approval': 'Sample Approval Email',
    'order_initiation': 'Order Initiation Email - 50% Order Booking',
    'packaging_confirmation': 'Packaging Confirmation Email',
    'packaging_payment_confirmation': 'Packaging Payment Confirmation Email',
    'printing_confirmation': 'Printing Confirmation Email',
    'printing_payment_confirmation': 'Printing Payment Confirmation Email',
    'final_order_shipment': 'Final Order Shipment Email with Final Invoice',
    'invoice': 'Invoice Email',
    'payment_confirmation': 'Payment Confirmation Email',
    'order_confirmation': 'Order Confirmation Email',
}

# ---------------------------------------------------------------------------
# Shared status helpers
# ---------------------------------------------------------------------------

_LEAD_STATUS_LABEL_TO_KEY = {label.lower(): key for key, label in Client.LEAD_STATUS_CHOICES}
_LEAD_SUB_STATUS_LABEL_TO_KEY = {label.lower(): key for key, label in Client.LEAD_SUB_STATUS_CHOICES}

TEMPLATE_COLUMNS = [
    'full_name',
    'phone_number',
    'email',
    'company_name',
    'city',
    'no_of_products',
    'planned_selling_price_range',
    'how_many_units_per_product',
    'physical_address',
    'gst_details',
    'lead_status',
    'sub_status',
]

LEAD_STATUS_HINT = (
    'Allowed: Initial Conversation | Product Requirement Captured | Proposal | Costing | '
    'Sample | Order | Production | Testing | Filling | Order Dispatch | Order Closed | On Hold | Lead Closed'
)
SUB_STATUS_HINT = (
    'Optional. Depends on Lead Status. E.g. for Sample: Formula Created, Sample Made, Approved …'
)

PHONE_RE = re.compile(r'\d{10}$')


class _SafeDict(dict):
    """Lets templates reference optional tokens (e.g. unused product slots) without KeyError."""
    def __missing__(self, key):
        return ''


def _parse_phone(raw) -> str | None:
    """Extract trailing 10-digit Indian mobile from any common format."""
    if raw is None:
        return None
    cleaned = re.sub(r'[^\d]', '', str(raw))
    if len(cleaned) >= 10:
        return cleaned[-10:]
    return None


def _parse_lead_status(raw) -> str:
    if not raw:
        return 'initial_conversation'
    raw_lower = str(raw).strip().lower()
    if raw_lower in dict(Client.LEAD_STATUS_CHOICES):
        return raw_lower
    return _LEAD_STATUS_LABEL_TO_KEY.get(raw_lower, 'initial_conversation')


def _parse_sub_status(raw, lead_status: str) -> str:
    if not raw:
        return ''
    raw_lower = str(raw).strip().lower()
    valid = Client.VALID_SUB_STATUSES.get(lead_status, [])
    if raw_lower in valid:
        return raw_lower
    matched = _LEAD_SUB_STATUS_LABEL_TO_KEY.get(raw_lower, '')
    return matched if matched in valid else ''


# ---------------------------------------------------------------------------
# ViewSet
# ---------------------------------------------------------------------------

class ClientViewSet(viewsets.ModelViewSet):
    serializer_class = ClientSerializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'phone_no'

    def get_queryset(self):
        qs = Client.objects.select_related('poc').all()
        params = self.request.query_params

        q = params.get('q')
        if q:
            qs = qs.filter(Q(name__icontains=q) | Q(phone_no__icontains=q))

        poc_id = params.get('poc')
        if poc_id:
            qs = qs.filter(poc__id=poc_id)

        lead_status = params.get('lead_status')
        if lead_status:
            qs = qs.filter(lead_status=lead_status)

        lead_bucket = params.get('lead_bucket')
        if lead_bucket:
            qs = qs.filter(lead_status__in=LEAD_BUCKETS.get(lead_bucket, []))

        created_after = params.get('created_after')
        if created_after:
            qs = qs.filter(created_at__date__gte=created_after)

        created_before = params.get('created_before')
        if created_before:
            qs = qs.filter(created_at__date__lte=created_before)

        return qs

    # ------------------------------------------------------------------
    # GET /api/clients/upload-template/
    # ------------------------------------------------------------------
    @action(detail=False, methods=['get'], url_path='upload-template')
    def download_template(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Client Upload'

        header_font = openpyxl.styles.Font(bold=True)
        header_fill = openpyxl.styles.PatternFill('solid', fgColor='FFF3CD')

        # Header row
        for col_idx, col_name in enumerate(TEMPLATE_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill

        # Column widths: name  phone email company city  noprod price  units  addr   gst    lead_status  sub_status
        widths = [       25,   20,   30,   25,     15,   14,    28,    22,    35,    20,    45,          45        ]
        for col_idx, width in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        # Hints in row 2
        hint_font = openpyxl.styles.Font(italic=True, color='888888')
        ls_col = TEMPLATE_COLUMNS.index('lead_status') + 1
        ws.cell(row=2, column=ls_col, value=LEAD_STATUS_HINT).font = hint_font
        ss_col = TEMPLATE_COLUMNS.index('sub_status') + 1
        ws.cell(row=2, column=ss_col, value=SUB_STATUS_HINT).font = hint_font
        noprod_col = TEMPLATE_COLUMNS.index('no_of_products') + 1
        ws.cell(row=2, column=noprod_col, value='Integer, e.g. 5').font = hint_font
        units_col = TEMPLATE_COLUMNS.index('how_many_units_per_product') + 1
        ws.cell(row=2, column=units_col, value='Integer, e.g. 1000').font = hint_font
        price_col = TEMPLATE_COLUMNS.index('planned_selling_price_range') + 1
        ws.cell(row=2, column=price_col, value='Free text, e.g. ₹100–₹500').font = hint_font

        # One example row
        ws.cell(row=3, column=1, value='John Doe')
        ws.cell(row=3, column=2, value='p:+919876543210')
        ws.cell(row=3, column=3, value='john@example.com')
        ws.cell(row=3, column=4, value='Acme Corp')
        ws.cell(row=3, column=5, value='Mumbai')
        ws.cell(row=3, column=6, value=3)
        ws.cell(row=3, column=7, value='₹200–₹500')
        ws.cell(row=3, column=8, value=1000)
        ws.cell(row=3, column=9, value='123 MG Road, Mumbai 400001')
        ws.cell(row=3, column=10, value='27AAPFU0939F1ZV')
        ws.cell(row=3, column=11, value='Initial Conversation')
        ws.cell(row=3, column=12, value='')

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="client_upload_template.xlsx"'
        return response

    # ------------------------------------------------------------------
    # POST /api/clients/bulk-upload/
    # ------------------------------------------------------------------
    @action(detail=False, methods=['post'], url_path='bulk-upload')
    def bulk_upload(self, request):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'detail': 'No file provided. Send an Excel file as "file".'}, status=400)

        # Load workbook
        try:
            wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
        except Exception:
            return Response({'detail': 'Could not read the file. Please upload a valid .xlsx file.'}, status=400)

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return Response({'detail': 'The file is empty.'}, status=400)

        # Detect header row — find column indices
        header_row = [str(c).strip().lower() if c is not None else '' for c in rows[0]]

        def col(name):
            try:
                return header_row.index(name)
            except ValueError:
                return None

        idx_name    = col('full_name')
        idx_phone   = col('phone_number')
        idx_email   = col('email')
        idx_company = col('company_name')
        idx_city    = col('city')
        idx_noprod  = col('no_of_products')
        idx_price   = col('planned_selling_price_range')
        idx_units   = col('how_many_units_per_product')
        idx_address = col('physical_address')
        idx_gst     = col('gst_details')
        idx_lead_status = col('lead_status')
        idx_sub_status  = col('sub_status')

        if idx_name is None or idx_phone is None:
            return Response(
                {'detail': 'Required columns "full_name" and "phone_number" not found in the header row.'},
                status=400,
            )

        def cell(row_data, idx):
            if idx is None or idx >= len(row_data):
                return None
            v = row_data[idx]
            return str(v).strip() if v is not None else None

        def _int_or_none(row_data, idx):
            v = cell(row_data, idx)
            if v is None:
                return None
            try:
                return int(float(v))
            except (ValueError, TypeError):
                return None

        created = []
        skipped = []

        # Parse all valid rows first, then do one EXISTS check and one bulk insert.
        pending = []  # list of (row_num, name, phone, kwargs, email_warning)
        seen_phones = set()  # deduplicate within the file itself

        for row_num, row_data in enumerate(rows[1:], start=2):
            if all(v is None or str(v).strip() == '' for v in row_data):
                continue

            name = cell(row_data, idx_name)
            phone_raw = cell(row_data, idx_phone)

            if not name:
                skipped.append({'row': row_num, 'name': '—', 'phone': phone_raw or '—',
                                 'reason': 'Missing full_name'})
                continue

            phone = _parse_phone(phone_raw)
            if not phone:
                skipped.append({'row': row_num, 'name': name, 'phone': phone_raw or '—',
                                 'reason': 'Could not extract a 10-digit phone number'})
                continue

            if phone in seen_phones:
                skipped.append({'row': row_num, 'name': name, 'phone': phone,
                                 'reason': 'Duplicate phone number within the uploaded file'})
                continue
            seen_phones.add(phone)

            email = cell(row_data, idx_email) or ''
            email_warning = None
            if email and not re.fullmatch(r'[^@\s]+@[^@\s]+\.[^@\s]+', email):
                email_warning = f'Invalid email "{email}" stored as-is'

            lead_status = _parse_lead_status(cell(row_data, idx_lead_status))
            pending.append((row_num, name, phone, {
                'phone_no': phone,
                'name': name,
                'email': email,
                'company_name': cell(row_data, idx_company) or '',
                'city': cell(row_data, idx_city) or '',
                'no_of_products': _int_or_none(row_data, idx_noprod),
                'planned_selling_price_range': cell(row_data, idx_price) or '',
                'how_many_units_per_product': _int_or_none(row_data, idx_units),
                'physical_address': cell(row_data, idx_address) or '',
                'gst_details': cell(row_data, idx_gst) or '',
                'lead_status': lead_status,
                'lead_sub_status': _parse_sub_status(cell(row_data, idx_sub_status), lead_status),
                'poc': request.user,
            }, email_warning))

        # Single query to find all phones that already exist
        pending_phones = [p[2] for p in pending]
        existing_phones = set(
            Client.objects.filter(phone_no__in=pending_phones).values_list('phone_no', flat=True)
        )

        to_create = []
        for row_num, name, phone, kwargs, email_warning in pending:
            if phone in existing_phones:
                skipped.append({'row': row_num, 'name': name, 'phone': phone,
                                 'reason': 'Phone number already exists in the system'})
                continue
            to_create.append((row_num, name, phone, kwargs, email_warning))

        # Single bulk insert (bulk_create bypasses auto_now so backfill updated_at)
        Client.objects.bulk_create([Client(**kw) for _, _, _, kw, _ in to_create])
        Client.objects.filter(
            phone_no__in=[p for _, _, p, _, _ in to_create]
        ).update(updated_at=timezone.now())

        for row_num, name, phone, _, email_warning in to_create:
            entry = {'row': row_num, 'name': name, 'phone': phone}
            if email_warning:
                entry['warning'] = email_warning
            created.append(entry)

        wb.close()
        return Response({'created': created, 'skipped': skipped}, status=200)

    # ------------------------------------------------------------------
    # POST /api/clients/send-welcome-email/
    # ------------------------------------------------------------------
    @action(detail=False, methods=['post'], url_path='send-welcome-email',
            parser_classes=[MultiPartParser, FormParser, JSONParser])
    def send_welcome_email(self, request):
        # Support both JSON (bulk, no files) and multipart (project email with files)
        if hasattr(request.data, 'getlist'):
            raw = request.data.get('phone_nos', '[]')
            phone_nos = json.loads(raw) if isinstance(raw, str) else raw
            email_type = request.data.get('email_type', 'welcome')
            project_id = request.data.get('project_id') or None
            uploaded_files = request.FILES.getlist('files')
            raw_ctx = request.data.get('extra_ctx', '{}')
            extra_ctx = json.loads(raw_ctx) if isinstance(raw_ctx, str) else {}
            raw_inv = request.data.get('invoice_ids', '[]')
            invoice_ids = json.loads(raw_inv) if isinstance(raw_inv, str) else raw_inv
        else:
            phone_nos = request.data.get('phone_nos', [])
            email_type = request.data.get('email_type', 'welcome')
            project_id = request.data.get('project_id') or None
            uploaded_files = []
            extra_ctx = request.data.get('extra_ctx', {})
            invoice_ids = request.data.get('invoice_ids', [])

        if not isinstance(phone_nos, list) or not phone_nos:
            return Response({'detail': 'phone_nos must be a non-empty list.'}, status=400)

        tpl = _TEMPLATE_MAP.get(email_type, welcome_tpl)

        # Upload attachments to Google Drive once (shared across recipients)
        drive_attachments = []
        if uploaded_files:
            from apps.files.drive_service import upload_file as drive_upload
            first_phone = str(phone_nos[0]).strip()
            try:
                first_client = Client.objects.get(phone_no=first_phone)
                client_name_for_drive = first_client.name
            except Client.DoesNotExist:
                client_name_for_drive = 'unknown'
            for f in uploaded_files:
                file_bytes = f.read()
                try:
                    result = drive_upload(
                        file_bytes=file_bytes,
                        filename=f.name,
                        mimetype=f.content_type or 'application/octet-stream',
                        client_name=client_name_for_drive,
                        subfolder='Emails',
                    )
                    drive_attachments.append({
                        'filename': f.name,
                        'drive_url': result['drive_url'],
                        'drive_file_id': result['drive_file_id'],
                        'content_type': f.content_type or 'application/octet-stream',
                        'bytes': file_bytes,
                    })
                except Exception as exc:
                    return Response({'detail': f'Drive upload failed for {f.name}: {exc}'}, status=502)

        # Attach selected generated invoices (download bytes from Drive)
        if invoice_ids:
            from apps.invoices.models import Invoice as InvoiceModel
            from apps.files.drive_service import download_file as drive_download
            for inv_id in invoice_ids:
                try:
                    inv = InvoiceModel.objects.get(pk=inv_id)
                    if inv.drive_file_id:
                        inv_bytes = drive_download(inv.drive_file_id)
                        drive_attachments.append({
                            'filename': f'{inv.invoice_number}.pdf',
                            'drive_url': inv.drive_url,
                            'drive_file_id': inv.drive_file_id,
                            'content_type': 'application/pdf',
                            'bytes': inv_bytes,
                        })
                except Exception:
                    pass  # skip individual invoice attachment failures silently

        # Resolve project FK once
        project_obj = None
        if project_id:
            from apps.crm_projects.models import CRMProject
            try:
                project_obj = CRMProject.objects.get(pk=project_id)
            except CRMProject.DoesNotExist:
                pass

        sent = []
        skipped = []
        sent_by_name = getattr(request.user, 'name', '') or request.user.email

        for phone in phone_nos:
            try:
                client = Client.objects.get(phone_no=str(phone).strip())
            except Client.DoesNotExist:
                skipped.append({'phone_no': phone, 'reason': 'Client not found'})
                continue

            if not client.email:
                skipped.append({'phone_no': phone, 'reason': 'No email address on file'})
                continue

            ctx = {
                'client_name': client.name,
                'company_name': client.company_name or '',
                'company_line': f' and {client.company_name}' if client.company_name else '',
                'sent_by_name': sent_by_name,
                **extra_ctx,
            }

            safe_ctx = _SafeDict(ctx)
            subject = tpl.SUBJECT.format_map(safe_ctx)
            html_body = tpl.HTML_BODY.format_map(safe_ctx)
            text_body = tpl.TEXT_BODY.format_map(safe_ctx)

            try:
                msg = EmailMultiAlternatives(subject=subject, body=text_body, to=[client.email])
                msg.attach_alternative(html_body, 'text/html')
                for att in drive_attachments:
                    msg.attach(att['filename'], att['bytes'], att['content_type'])
                msg.send()

                EmailLog.objects.create(
                    client=client,
                    project=project_obj,
                    email_type=email_type,
                    email_type_label=_TEMPLATE_LABELS.get(email_type, email_type),
                    recipient_email=client.email,
                    subject=subject,
                    sent_by=request.user,
                    sent_by_name=sent_by_name,
                    attachments=[
                        {'filename': a['filename'], 'drive_url': a['drive_url'], 'drive_file_id': a['drive_file_id']}
                        for a in drive_attachments
                    ],
                )
                sent.append(phone)
            except Exception as exc:
                skipped.append({'phone_no': phone, 'reason': f'Send failed: {exc}'})

        return Response({'sent': sent, 'skipped': skipped}, status=200)

    @action(detail=True, methods=['get'], url_path='email-history')
    def email_history(self, request, phone_no=None):
        logs = (
            EmailLog.objects
            .filter(client_id=phone_no)
            .select_related('project', 'sent_by')
            .order_by('-sent_at')
        )
        return Response(EmailLogSerializer(logs, many=True).data)

    # ------------------------------------------------------------------
    # GET /api/clients/lead-bucket-counts/
    # ------------------------------------------------------------------
    @action(detail=False, methods=['get'], url_path='lead-bucket-counts')
    def lead_bucket_counts(self, request):
        qs = self.get_queryset()
        counts = {
            bucket: qs.filter(lead_status__in=statuses).count()
            for bucket, statuses in LEAD_BUCKETS.items()
        }
        return Response(counts)

    # ------------------------------------------------------------------
    # PATCH /api/clients/bulk-update-status/
    # ------------------------------------------------------------------
    @action(detail=False, methods=['patch'], url_path='bulk-update-status')
    def bulk_update_status(self, request):
        phone_nos = request.data.get('phone_nos', [])
        new_status = request.data.get('lead_status', '')
        new_sub_status = request.data.get('lead_sub_status', '')

        if not isinstance(phone_nos, list) or not phone_nos:
            return Response({'detail': 'phone_nos must be a non-empty list.'}, status=400)

        valid_statuses = dict(Client.LEAD_STATUS_CHOICES)
        if new_status not in valid_statuses:
            return Response(
                {'detail': f'Invalid lead_status "{new_status}". Valid values: {list(valid_statuses.keys())}'},
                status=400,
            )

        # Validate sub_status if provided
        if new_sub_status:
            valid_sub = Client.VALID_SUB_STATUSES.get(new_status, [])
            if new_sub_status not in valid_sub:
                return Response(
                    {'detail': f'Invalid lead_sub_status "{new_sub_status}" for status "{new_status}".'},
                    status=400,
                )

        fields = {'lead_status': new_status, 'lead_sub_status': new_sub_status}
        updated = Client.objects.filter(phone_no__in=phone_nos).update(**fields)
        return Response({'updated': updated}, status=200)
