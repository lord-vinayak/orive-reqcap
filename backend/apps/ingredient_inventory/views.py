import io
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

import openpyxl
from django.db.models import Q
from django.http import HttpResponse
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from .models import IngredientRecord
from .serializers import IngredientRecordSerializer

TEMPLATE_COLUMNS = [
    'date', 'category', 'client_name', 'vendor_name', 'product_name',
    'ingredient_use', 'quantity_sent', 'used_in_batch', 'comment',
]

_DATE_FORMATS = [
    '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y',
    '%b %Y', '%B %Y', '%b-%Y', '%B-%Y',  # month + year only, e.g. "Feb 2026" — defaults to the 1st
]


def _parse_date(raw):
    if raw is None or raw == '':
        return None
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw
    raw = str(raw).strip()
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def _decimal_or_none(raw):
    if raw is None or str(raw).strip() == '':
        return None
    try:
        return Decimal(str(raw).strip())
    except InvalidOperation:
        return None


class IngredientRecordViewSet(viewsets.ModelViewSet):
    serializer_class = IngredientRecordSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = IngredientRecord.objects.all()
        q = self.request.query_params.get('q')
        if q:
            qs = qs.filter(
                Q(client_name__icontains=q) | Q(vendor_name__icontains=q) |
                Q(product_name__icontains=q) | Q(category__icontains=q)
            )
        return qs

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    def destroy(self, request, *args, **kwargs):
        if request.user.role != 'admin':
            return Response({'detail': 'Only admin can delete.'}, status=403)
        return super().destroy(request, *args, **kwargs)

    # ------------------------------------------------------------------
    # GET /api/ingredient-records/upload-template/
    # ------------------------------------------------------------------
    @action(detail=False, methods=['get'], url_path='upload-template')
    def download_template(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Ingredient Inventory'

        header_font = openpyxl.styles.Font(bold=True)
        header_fill = openpyxl.styles.PatternFill('solid', fgColor='FFF3CD')

        for col_idx, col_name in enumerate(TEMPLATE_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill

        widths = [14, 20, 22, 22, 26, 30, 14, 14, 30]
        for col_idx, width in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

        hint_font = openpyxl.styles.Font(italic=True, color='888888')
        date_col = TEMPLATE_COLUMNS.index('date') + 1
        ws.cell(row=2, column=date_col, value='YYYY-MM-DD, DD-MM-YYYY, or month + year e.g. Feb 2026').font = hint_font

        ws.cell(row=3, column=1, value='2026-06-01')
        ws.cell(row=3, column=2, value='Raw Material- Actives')
        ws.cell(row=3, column=3, value='Shubham')
        ws.cell(row=3, column=4, value='Sahil Caldic')
        ws.cell(row=3, column=5, value='MaizeCare Style')
        ws.cell(row=3, column=6, value='Sample for stability testing')
        ws.cell(row=3, column=7, value=100)
        ws.cell(row=3, column=8, value=5)
        ws.cell(row=3, column=9, value='From raw material sample')

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="ingredient_inventory_template.xlsx"'
        return response

    # ------------------------------------------------------------------
    # POST /api/ingredient-records/bulk-upload/
    # ------------------------------------------------------------------
    @action(detail=False, methods=['post'], url_path='bulk-upload')
    def bulk_upload(self, request):
        file_obj = request.FILES.get('file')
        if not file_obj:
            return Response({'detail': 'No file provided. Send an Excel file as "file".'}, status=400)

        try:
            wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
        except Exception:
            return Response({'detail': 'Could not read the file. Please upload a valid .xlsx file.'}, status=400)

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return Response({'detail': 'The file is empty.'}, status=400)

        header_row = [str(c).strip().lower() if c is not None else '' for c in rows[0]]

        def col(name):
            try:
                return header_row.index(name)
            except ValueError:
                return None

        idx = {name: col(name) for name in TEMPLATE_COLUMNS}

        if idx['client_name'] is None or idx['vendor_name'] is None or idx['product_name'] is None:
            return Response(
                {'detail': 'Required columns "client_name", "vendor_name" and "product_name" not found in the header row.'},
                status=400,
            )

        def cell(row_data, name):
            i = idx[name]
            if i is None or i >= len(row_data):
                return None
            v = row_data[i]
            return str(v).strip() if v is not None else None

        created = []
        skipped = []
        to_create = []

        for row_num, row_data in enumerate(rows[1:], start=2):
            if all(v is None or str(v).strip() == '' for v in row_data):
                continue

            client_name = cell(row_data, 'client_name')
            vendor_name = cell(row_data, 'vendor_name')
            product_name = cell(row_data, 'product_name')

            if not client_name or not vendor_name or not product_name:
                skipped.append({
                    'row': row_num,
                    'client_name': client_name or '—',
                    'vendor_name': vendor_name or '—',
                    'reason': 'Missing client_name, vendor_name, or product_name',
                })
                continue

            date_raw = row_data[idx['date']] if idx['date'] is not None and idx['date'] < len(row_data) else None
            parsed_date = _parse_date(date_raw)

            to_create.append(IngredientRecord(
                date=parsed_date,
                category=cell(row_data, 'category') or '',
                client_name=client_name,
                vendor_name=vendor_name,
                product_name=product_name,
                ingredient_use=cell(row_data, 'ingredient_use') or '',
                quantity_sent=_decimal_or_none(cell(row_data, 'quantity_sent')),
                used_in_batch=_decimal_or_none(cell(row_data, 'used_in_batch')),
                comment=cell(row_data, 'comment') or '',
                created_by=request.user,
            ))
            entry = {'row': row_num, 'client_name': client_name, 'vendor_name': vendor_name}
            if date_raw and not parsed_date:
                entry['warning'] = f'Could not parse date "{date_raw}" — left blank'
            created.append(entry)

        IngredientRecord.objects.bulk_create(to_create)
        wb.close()
        return Response({'created': created, 'skipped': skipped}, status=200)
