"""Build a formatted XLSX Client Costing.

Reads from each item's merged catalog/snapshot view so user-edits to costing cells
are honored. Also embeds the company logo (frontend/public/logo.png) at the
top-left of the sheet (item #2).
"""
import io
import os
import re
from datetime import date
from pathlib import Path
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlsxImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


MUSTARD = 'D6AE4D'
LIGHT_GREY = 'FFFCF3'
BLACK = '000000'

THIN = Side(border_style='thin', color='999999')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# Resolve the logo file at backend/.. /frontend/public/logo.png
# backend/apps/proposals/xlsx_export.py  ->  go up 3, then frontend/public/logo.png
_HERE = Path(__file__).resolve()
_LOGO_CANDIDATES = [
    _HERE.parents[2] / 'logo.png',                          # bundled in Docker image (/app/logo.png)
    _HERE.parents[3] / 'frontend' / 'public' / 'logo.png', # local dev (source tree)
    _HERE.parents[3] / 'frontend' / 'dist' / 'logo.png',   # local dev (built)
]


def _find_logo():
    for p in _LOGO_CANDIDATES:
        if p.exists():
            return str(p)
    # Fall back to env var override if deployment has a different layout.
    env_path = os.environ.get('SKINOVATION_LOGO_PATH')
    if env_path and Path(env_path).exists():
        return env_path
    return None


def _merged(item):
    """Return a dict of effective field values: catalog defaults overlaid by snapshot."""
    out = {}
    c = item.catalog_item
    if c:
        for f in [
            'body_part', 'product_type', 'sub_product_type',
            'kb_tag1', 'kb_tag2', 'kb_tag3',
            'specific_ingredients', 'color', 'fragrance', 'size', 'packaging_type',
            'rate_category', 'per_kg_rate', 'manufacturing_cost', 'rate_per_unit',
            'tentative_packaging_cost', 'label_cost', 'tentative_monocarton_cost',
            'total_cost', 'potential_mrp',
        ]:
            out[f] = getattr(c, f, None)
    for k, v in (item.snapshot or {}).items():
        if v is not None and v != '':
            out[k] = v
    return out


def _dec(val):
    """Coerce a numeric/string/decimal/None into float-or-empty for an Excel cell."""
    if val is None or val == '':
        return ''
    try:
        return float(val)
    except (TypeError, ValueError):
        return val


def _compute_costs(m):
    """Derive the four calculated cost columns from merged row data.

    Returns (raw_material_per_unit, estimated_unit_cost, total_cost, potential_mrp).
    Any value that cannot be computed is returned as None.
    """
    def _n(k):
        v = m.get(k)
        if v is None or v == '':
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    per_kg = _n('per_kg_rate')
    # Size may include units like "100g" or "500ml" — extract the leading number.
    size_str = str(m.get('size') or '')
    size_match = re.match(r'[\d.]+', size_str)
    try:
        size_val = float(size_match.group()) if size_match else None
    except (TypeError, ValueError):
        size_val = None

    mfg = _n('manufacturing_cost')
    pkg = _n('tentative_packaging_cost')
    label = _n('label_cost')
    mono = _n('tentative_monocarton_cost')

    raw_per_unit = (per_kg / 1000 * size_val) if (per_kg is not None and size_val is not None) else None
    est_unit = (raw_per_unit + mfg) if (raw_per_unit is not None and mfg is not None) else None
    total = (est_unit + (pkg or 0) + (label or 0) + (mono or 0)) if est_unit is not None else None
    mrp = (total * 6) if total is not None else None

    return raw_per_unit, est_unit, total, mrp


def build_proposal_xlsx(proposal):
    """Return bytes of a formatted .xlsx for the given Client Costing."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Client Costing'

    req = proposal.requirement
    client = req.client
    items = list(proposal.items.select_related('catalog_item').order_by('sort_order'))

    # Row 1: Logo (top-left) + company name banner.
    # We leave A1 small for the logo image and put the title in B1:N1.
    ws.row_dimensions[1].height = 54

    logo_path = _find_logo()
    if logo_path:
        try:
            img = XlsxImage(logo_path)
            # Resize so it fits the header row.
            img.width = 70
            img.height = 70
            ws.add_image(img, 'A1')
        except Exception:
            # If openpyxl can't embed the image (e.g. corrupt), continue without it.
            pass

    ws['A1'].fill = PatternFill('solid', fgColor=MUSTARD)

    # Rows 3-6: Date + client info
    ws['A3'] = 'Date'
    ws['B3'] = date.today().isoformat()
    ws['A4'] = 'Client Name'
    ws['B4'] = client.name
    ws['A5'] = 'Phone'
    ws['B5'] = client.phone_no
    ws['A6'] = 'Point of Contact'
    ws['B6'] = client.poc.name if client.poc else ''
    for r in range(3, 7):
        ws.cell(row=r, column=1).font = Font(name='Aptos', bold=True)
        ws.cell(row=r, column=2).font = Font(name='Aptos')

    # Row 8: Column headers
    header_row = 8
    columns = [
        ('Body Part',                        'body_part'),
        ('Product Type',                     'product_type'),
        ('Sub Product Type',                 'sub_product_type'),
        ('Key Benefits',                     '__kb__'),
        ('Specific Ingredients',             'specific_ingredients'),
        ('Color',                            'color'),
        ('Fragrance',                        'fragrance'),
        ('Size',                             'size'),
        ('Packaging',                        'packaging_type'),
        ('Raw Material Cost (per kg)',       'per_kg_rate'),
        ('Raw Material Cost (per unit)',     '__raw_per_unit__'),
        ('Manufacturing Cost',               'manufacturing_cost'),
        ('Estimated Unit Cost',              '__est_unit__'),
        ('Tentative Packaging Cost ~',        'tentative_packaging_cost'),
        ('Label Cost ~',                     'label_cost'),
        ('Tentative Monocarton Cost ~',      'tentative_monocarton_cost'),
        ('Total Cost ~',                     '__total__'),
        ('Potential MRP ~',                  '__mrp__'),
    ]

    # Now that we know the column count, apply the banner merges dynamically
    # so they always stretch across the full table width.
    last_col_letter = get_column_letter(len(columns))

    ws.merge_cells(f'B1:{last_col_letter}1')
    ws['B1'] = 'SKINOVATION SCIENCES'
    ws['B1'].font = Font(name='Aptos', size=18, bold=True, color=BLACK)
    ws['B1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B1'].fill = PatternFill('solid', fgColor=MUSTARD)

    ws.merge_cells(f'A2:{last_col_letter}2')
    ws['A2'] = 'Client Costing'
    ws['A2'].font = Font(name='Aptos', size=14, bold=True, color=BLACK)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A2'].fill = PatternFill('solid', fgColor=LIGHT_GREY)
    ws.row_dimensions[2].height = 28

    for i, (label, _) in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=i, value=label)
        cell.font = Font(name='Aptos', bold=True, color=BLACK)
        cell.fill = PatternFill('solid', fgColor=MUSTARD)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[header_row].height = 28

    # Data rows — pull from merged catalog+snapshot view.
    for idx, item in enumerate(items, start=1):
        m = _merged(item)
        kb = ', '.join(str(m.get(k) or '') for k in ('kb_tag1', 'kb_tag2', 'kb_tag3')
                       if m.get(k))
        raw_per_unit, est_unit, total_cost, potential_mrp = _compute_costs(m)
        values = [
            m.get('body_part') or '',
            m.get('product_type') or '',
            m.get('sub_product_type') or '',
            kb,
            m.get('specific_ingredients') or '',
            m.get('color') or '',
            m.get('fragrance') or '',
            m.get('size') or '',
            m.get('packaging_type') or '',
            _dec(m.get('per_kg_rate')),
            raw_per_unit if raw_per_unit is not None else '',
            _dec(m.get('manufacturing_cost')),
            est_unit if est_unit is not None else '',
            _dec(m.get('tentative_packaging_cost')),
            _dec(m.get('label_cost')),
            _dec(m.get('tentative_monocarton_cost')),
            total_cost if total_cost is not None else '',
            potential_mrp if potential_mrp is not None else '',
        ]
        for i, val in enumerate(values, start=1):
            cell = ws.cell(row=header_row + idx, column=i, value=val)
            cell.font = Font(name='Aptos')
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.border = BORDER
            if idx % 2 == 0:
                cell.fill = PatternFill('solid', fgColor=LIGHT_GREY)

    footnote_row = header_row + len(items) + 1
    ws.merge_cells(f'A{footnote_row}:{last_col_letter}{footnote_row}')
    fn_cell = ws.cell(row=footnote_row, column=1, value='~ Tentative figures — subject to revision')
    fn_cell.font = Font(name='Aptos', italic=True, color='806600')
    fn_cell.alignment = Alignment(horizontal='left', vertical='center')

    widths = [14, 16, 18, 26, 32, 12, 12, 10, 14, 22, 22, 18, 20, 22, 14, 22, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
