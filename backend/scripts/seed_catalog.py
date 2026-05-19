"""
Seed CatalogItem from a local .xlsx file.

Usage (inside backend/ with venv active):
    python scripts/seed_catalog.py "E:/Orive/Content for Website.xlsx"

Re-runs are safe: existing rows with the same (client_name, body_part,
product_type, sub_product_type, size) signature are not duplicated.
"""
import sys
import os
import django

# Set up Django
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'core.settings')
django.setup()

from openpyxl import load_workbook
from apps.catalog.models import CatalogItem


COLUMN_MAP = {
    'date': 'date', 'poc': 'poc', 'client name': 'client_name',
    'body part': 'body_part', 'product type': 'product_type',
    'sub product type': 'sub_product_type',
    'kb tag1': 'kb_tag1', 'kb tag2': 'kb_tag2', 'kb tag3': 'kb_tag3',
    'specific ingredients': 'specific_ingredients',
    'color': 'color', 'fragrance': 'fragrance',
    'size': 'size', 'packaging type': 'packaging_type',
    'per kg rate': 'per_kg_rate', 'manufacturing cost': 'manufacturing_cost',
    'rate per unit': 'rate_per_unit',
    'tentative packaging cost': 'tentative_packaging_cost',
    'label cost': 'label_cost',
    'tentative monocarton cost': 'tentative_monocarton_cost',
    'total cost': 'total_cost', 'potential mrp': 'potential_mrp',
}
DECIMAL_FIELDS = {
    'per_kg_rate', 'manufacturing_cost', 'rate_per_unit',
    'tentative_packaging_cost', 'label_cost',
    'tentative_monocarton_cost', 'total_cost', 'potential_mrp',
}


def main(path):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        print('Sheet is empty')
        return

    header_idx = 1 if rows[0][0] is None else 0
    header = [str(h).strip().lower() if h else '' for h in rows[header_idx]]
    data_rows = rows[header_idx + 1:]

    created = 0
    skipped = 0
    for raw in data_rows:
        if not raw or all(v is None for v in raw):
            continue
        obj = {}
        for i, value in enumerate(raw):
            if i >= len(header):
                break
            field = COLUMN_MAP.get(header[i])
            if not field or value is None:
                continue
            if field in DECIMAL_FIELDS:
                try:
                    obj[field] = float(value)
                except (ValueError, TypeError):
                    continue
            else:
                obj[field] = str(value).strip()
        if not obj:
            continue

        sig = {
            'client_name': obj.get('client_name', ''),
            'body_part': obj.get('body_part', ''),
            'product_type': obj.get('product_type', ''),
            'sub_product_type': obj.get('sub_product_type', ''),
            'size': obj.get('size', ''),
        }
        if CatalogItem.objects.filter(**sig).exists():
            skipped += 1
            continue
        CatalogItem.objects.create(**obj)
        created += 1

    print(f'Created: {created} | Skipped (already present): {skipped}')


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('Usage: python scripts/seed_catalog.py <path/to/catalog.xlsx>')
        sys.exit(1)
    main(sys.argv[1])
